
import openpyxl, os.path, pprint

folder = 'D:\\pythonTestFiles'
filename = 'example.xlsx'

file1 = openpyxl.load_workbook(os.path.join(folder, filename))

print (file1.get_sheet_names())

sheet = file1.get_sheet_by_name('Лист1')

c = sheet['B1']
d = sheet['C1']

print ('A1 value is %s and B1 value is %s, C1 is %s'% (sheet['A1'].value, c.value, d.value))


for i in range(1,8):
    print (i, sheet.cell(row=i, column=2).value)

print (sheet.max_row)
print (sheet.max_column)


pprint.pprint (tuple(sheet['A1':'C3']))

for rawCellObj in sheet['A1':'C8']:
    for CellObj in rawCellObj:
        print (CellObj.coordinate, CellObj.value)
    print ('----------END OF ROW----------')


